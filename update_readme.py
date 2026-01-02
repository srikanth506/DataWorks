import pandas as pd
from openpyxl import load_workbook
import re
from datetime import datetime

EXCEL_FILE = "/Users/srikanth/Library/CloudStorage/OneDrive-Personal/DataWorks/DataWorks Progress Plan.xlsx"
README_FILE = "README.md"

START_MARKER = "<!-- AUTO-GENERATED-START -->"
END_MARKER = "<!-- AUTO-GENERATED-END -->"

MONTH_ORDER = {
    "January": 1, "February": 2, "March": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12
}


def parse_sheet_order(sheet_name):
    """Convert '2026 - January' → sortable (year, month_number)"""
    match = re.match(r"(\d{4})\s*-\s*(\w+)", sheet_name)
    if not match:
        return (9999, 99)
    year, month = match.groups()
    return (int(year), MONTH_ORDER.get(month, 99))


def read_sheet_with_links(sheet_name):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]

    headers = [cell.value for cell in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2):
        row_data = {}
        for header, cell in zip(headers, row):
            if cell.hyperlink:
                row_data[header] = f"[{cell.value}]({cell.hyperlink.target})"
            else:
                row_data[header] = cell.value
        rows.append(row_data)

    return pd.DataFrame(rows)


def df_to_markdown(df):
    df = df.fillna("")
    return df.to_markdown(index=False)


def generate_summary(df):
    df = df.fillna("").astype(str)

    # Remove completely empty rows
    df = df[~(df == "").all(axis=1)]

    return {
        "total_days": len(df),
        "sql": df["SQL"].str.strip().ne("").sum(),
        "big_data": df["Big Data"].str.strip().ne("").sum(),
        "ds": df["Data Science"].str.strip().ne("").sum(),
        "job": df["Job Search"].str.strip().ne("").sum(),
    }


def build_auto_section():
    wb = load_workbook(EXCEL_FILE)
    sheets = sorted(wb.sheetnames, key=parse_sheet_order)

    all_data = []
    month_sections = []

    for sheet in sheets:
        df = read_sheet_with_links(sheet)
        all_data.append(df)

        month_sections.append(
            f"### 📅 {sheet}\n\n{df_to_markdown(df)}\n"
        )

    full_df = pd.concat(all_data, ignore_index=True)
    summary = generate_summary(full_df)

    auto_content = f"""
# DataWorks
I will use this repository to add the files related to big data, data science, data analytics projects


## 📊 Progress Summary

- **Total Days Logged:** {summary['total_days']}
- **SQL Topics Covered:** {summary['sql']} days
- **Big Data Activities:** {summary['big_data']} days
- **Data Science Activities:** {summary['ds']} days
- **Job Search Activities:** {summary['job']} days


## Career Progress Plan

This document tracks my daily and weekly progress in SQL, Big Data, Data Science, and Job Search activities.

{chr(10).join(month_sections)}
"""
    return auto_content.strip()


def update_readme():
    auto_section = f"{START_MARKER}\n{build_auto_section()}\n{END_MARKER}"

    try:
        with open(README_FILE, "r", encoding="utf-8") as f:
            content = f.read()

        if START_MARKER in content and END_MARKER in content:
            content = re.sub(
                f"{START_MARKER}.*?{END_MARKER}",
                auto_section,
                content,
                flags=re.S
            )
        else:
            content = auto_section

    except FileNotFoundError:
        content = auto_section

    with open(README_FILE, "w", encoding="utf-8") as f:
        f.write(content)


if __name__ == "__main__":
    update_readme()
